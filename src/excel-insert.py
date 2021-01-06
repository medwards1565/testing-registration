'''
Created on Dec 22, 2020

@author: medwa
'''

import openpyxl

qrcode = 'michael,edwards,1986-04-26,162222c615e70,therookie'

# xlfile is path to .xlsx file, will eventually be read from scanner
# tempsheet is name of sheet that will be created and manipulated by program
# mastersheet is name of sheet that needs to be parsed
def copy_to_new_sheet(xlfile, tempsheet, mastersheet):
    # load workbook
    print('test-registration(copy_to_new_sheet): Loading workbook ' + xlfile)
    wb = openpyxl.load_workbook(xlfile)
    
    # create new sheet
    print('test-registration(copy_to_new_sheet): Creating new sheet ' + tempsheet + ' in ' + xlfile)
    wb.create_sheet(tempsheet)
    
    # get the sheet to load data from
    temp_sheet = wb.get_sheet_by_name(tempsheet)
    
    # copy data
    print('test-registration(copy_to_new_sheet): Copying data from ' + mastersheet + ' to ' + tempsheet + ' in ' + xlfile)
    for row in wb [mastersheet] ['B3':'H14']:
        for cell in row:
            temp_sheet[cell.coordinate] = cell.value
    
    # save the workbook
    wb.save('C:\\Users\\medwa\\Desktop\\xlfolder\\test_copy.xlsx')

def match_and_log_user(qrcode, barcode, xlfile, sheet):
    # split qrcode string
    userinfo = qrcode.split(',')
    
    # load workbook
    print('test-registration(match_and_log_user): Loading workbook ' + xlfile)
    wb = openpyxl.load_workbook(xlfile)
    
    # get the sheet to match against
    sheetobj = wb.get_sheet_by_name(sheet)
    
    # search sheet for matching userinfo
    # make list of all users, then if comparison?
    
    
# function calls
# copy_to_new_sheet('C:\\Users\\medwa\\Desktop\\xlfolder\\test.xlsx', 'TEMPSHEET', 'Sheet1')
match_and_log_user(qrcode, 'BR12345678', 'C:\\Users\\medwa\\Desktop\\xlfolder\\test.xlsx', 'TEMPSHEET')